"""Allegro pin 表 loader（xls/xlsx）+ 筛选文件（默认 .lst）。

**输入 1 —— Allegro 导出的表格**（xls/xlsx），列含：
`REFDES, PIN_NUMBER, SYM_NAME, COMP_DEVICE_TYPE, PAD_STACK_NAME, PIN_X, PIN_Y, NET_NAME`（单位 mm）。

**输入 2 —— 筛选文件**（暂定方案）：
- 默认 `.lst`/`.txt`：纯文本，**一行一个 net 名**，空行和 `#` 注释行跳过；
- 兼容 `.xls`/`.xlsx`：表格，第一列是 net 名。
不在筛选文件里的 net **全部不要**（以后继续优化）。
"""
from __future__ import annotations

import os
import re
from collections import defaultdict

from ..model import (Point, Units, LayerDef, LayerStack, SignalGroup,
                     NetClass, Net, Pin)
from .loader import LoadedData
from .wire_gen import generate_wires

# 列名别名：表头（小写）→ 规范列名
_COL_ALIASES = {
    "net": "net_name", "net_name": "net_name", "network": "net_name", "网络": "net_name",
    "pin_x": "pin_x", "x": "pin_x", "x坐标": "pin_x",
    "pin_y": "pin_y", "y": "pin_y", "y坐标": "pin_y",
    "refdes": "refdes", "reference": "refdes", "位号": "refdes",
    "pin_number": "pin_number", "pin": "pin_number", "引脚": "pin_number",
}

_HEADER_WORDS = {"net", "net_name", "netname", "network", "名称", "网络", "net list"}


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_rows(path: str):
    """逐行 yield xls/xlsx 表格内容（tuple）。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(values_only=True):
                yield row
        finally:
            wb.close()
    elif ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            yield tuple(ws.row_values(r))
    else:
        raise ValueError(f"不支持的表格格式: {ext!r}（仅支持 .xls/.xlsx）")


def classify_net(name: str) -> NetClass | None:
    """net 名 → 类别；NC 返回 None（直接删）。"""
    u = name.upper().strip()
    if u == "NC":
        return None
    if any(k in u for k in ("GND", "VSS", "AGND", "DGND", "SGND", "GROUND")):
        return NetClass.GROUND
    if any(k in u for k in ("VDD", "VCC", "VPP", "VREF", "AVDD", "DVDD", "PWR")):
        return NetClass.POWER
    if re.search(r"V\d", u):
        return NetClass.POWER
    return NetClass.SIGNAL


def _read_text_net_list(path: str) -> set[str]:
    """读纯文本筛选文件（.lst/.txt）：一行一个 net，跳过空行和 # 注释。"""
    names: set[str] = set()
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            v = line.strip()
            if not v or v.startswith("#"):
                continue
            names.add(v)
    return names


def read_net_filter(path: str) -> set[str]:
    """读筛选文件：.lst/.txt 按文本（一行一个 net）；.xls/.xlsx 按表格第一列。"""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".lst", ".txt"):
        return _read_text_net_list(path)
    return _read_net_whitelist_table(path)


def _read_net_whitelist_table(path: str) -> set[str]:
    """读筛选表格（xls/xlsx）：第一列是 net 名（跳过表头行、空行）。"""
    names: set[str] = set()
    for row in _read_rows(path):
        if not row or row[0] is None:
            continue
        v = _cell_str(row[0])
        if not v:
            continue
        if v.lower() in _HEADER_WORDS:
            continue
        names.add(v)
    return names


class XlsxLoader:
    """读 pin 表 → net 聚合 → 分类 / 白名单筛选 → 飞线生成。"""

    def __init__(self, filter_path: str | None = None, n_signal_layers: int = 4,
                 width: float = 0.2, clearance: float = 0.2):
        self.filter_path = filter_path
        self.n_signal_layers = n_signal_layers
        self.width = width
        self.clearance = clearance

    def load(self, path: str) -> LoadedData:
        warnings: list[str] = []
        whitelist = read_net_filter(self.filter_path) if self.filter_path else None

        rows = list(_read_rows(path))
        if not rows:
            raise ValueError(f"表格为空: {path}")
        net_i, x_i, y_i, ref_i, pin_i = self._columns(rows[0])

        net_pins: dict[str, list[Pin]] = defaultdict(list)
        counter: dict[str, int] = defaultdict(int)
        for row in rows[1:]:
            if len(row) <= max(net_i, x_i, y_i):
                continue
            net = _cell_str(row[net_i]) if row[net_i] is not None else ""
            if not net:
                continue
            try:
                x = float(row[x_i])
                y = float(row[y_i])
            except (TypeError, ValueError):
                continue
            ref = _cell_str(row[ref_i]) if ref_i is not None and len(row) > ref_i else ""
            pin = _cell_str(row[pin_i]) if pin_i is not None and len(row) > pin_i else ""
            pid = f"{ref}.{pin}" if ref and pin else f"{net}.{counter[net]}"
            counter[net] += 1
            net_pins[net].append(Pin(pid, Point(x, y)))

        stack = LayerStack(
            tuple(LayerDef(i, f"L{i}", "signal", "any")
                  for i in range(1, self.n_signal_layers + 1)),
            via_kind="through")

        nets: list[Net] = []
        for net, pins in sorted(net_pins.items()):
            nc = classify_net(net)
            if nc is None:                     # NC 直接删
                continue
            if whitelist is not None and net not in whitelist:   # 白名单外全部不要
                continue
            if len(pins) < 2:                  # 单 pin 无法成飞线
                continue
            nets.append(Net(net, nc, None, None, tuple(pins),
                            self.width, self.clearance))
        if whitelist is not None:
            warnings.append(f"白名单筛选：保留 {len(nets)} 个 net（原始 {len(net_pins)} 个）")

        sig_nets = tuple(n for n in nets if n.net_class == NetClass.SIGNAL)
        groups = (SignalGroup("default", tuple(stack.signal_layers()),
                              tuple(n.net_id for n in sig_nets)),)
        wires = generate_wires(tuple(nets), warnings)
        return LoadedData(stack, groups, (), tuple(nets), (), wires, Units.MM,
                          tuple(warnings))

    @staticmethod
    def _columns(header_row) -> tuple[int, int, int, int | None, int | None]:
        col: dict[str, int] = {}
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            canonical = _COL_ALIASES.get(_cell_str(cell).lower())
            if canonical and canonical not in col:
                col[canonical] = i
        if "net_name" in col and "pin_x" in col and "pin_y" in col:
            return (col["net_name"], col["pin_x"], col["pin_y"],
                    col.get("refdes"), col.get("pin_number"))
        # 无表头 / 表头不识别 → 按 Allegro 导出固定列序
        return (7, 5, 6, 0, 1)
